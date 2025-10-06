"""Inject the standardized CMO/resin Policy sheet into an Excel workbook."""

from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName

_POLICY_VALUE_LINKS = {
    "Retainer_Fee_per_Year": "CMO_Retainer_Fee_per_Year",
    "CIP_Cost_per_Cycle": "Chromatography_CIP_Cost_per_Cycle",
}


def _copy_policy_sheet(source_sheet, target_workbook) -> None:
    if "Policy" in target_workbook.sheetnames:
        del target_workbook["Policy"]
    policy_sheet = target_workbook.create_sheet("Policy")

    for row in source_sheet.iter_rows():
        for cell in row:
            target_cell = policy_sheet.cell(row=cell.row, column=cell.column)
            target_cell.value = cell.value
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.fill = copy(cell.fill)
                target_cell.border = copy(cell.border)
                target_cell.number_format = cell.number_format
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)

    for col_letter, dimension in source_sheet.column_dimensions.items():
        target_dimension = policy_sheet.column_dimensions[col_letter]
        target_dimension.width = dimension.width

    for idx, dimension in source_sheet.row_dimensions.items():
        policy_sheet.row_dimensions[idx].height = dimension.height

    for merged in source_sheet.merged_cells.ranges:
        policy_sheet.merge_cells(str(merged))


def _link_policy_values(policy_sheet, target_workbook) -> None:
    """Point Policy value cells at workbook-level named ranges when available."""

    if not _POLICY_VALUE_LINKS:
        return

    defined_names = target_workbook.defined_names
    for row in policy_sheet.iter_rows(min_row=2, max_col=4):
        if len(row) < 4:
            continue
        parameter = row[1].value
        if parameter not in _POLICY_VALUE_LINKS:
            continue
        defined_name = _POLICY_VALUE_LINKS[parameter]
        if defined_name not in defined_names:
            continue
        cell = row[3]
        cell.value = f"={defined_name}"


def _copy_defined_names(source_workbook, target_workbook) -> None:
    for name in list(target_workbook.defined_names.keys()):
        if name in source_workbook.defined_names:
            del target_workbook.defined_names[name]

    for name, defined in source_workbook.defined_names.items():
        if not defined.attr_text:
            continue
        target_workbook.defined_names.add(
            DefinedName(name=name, attr_text=defined.attr_text)
        )


def inject_policy(target: Path, source: Path, *, backup: bool = True) -> Path:
    if backup:
        backup_path = target.with_name(f"{target.stem}_BACKUP{target.suffix}")
        shutil.copyfile(target, backup_path)
    else:
        backup_path = target

    source_wb = load_workbook(source, data_only=False)
    target_wb = load_workbook(target, data_only=False)
    try:
        _copy_policy_sheet(source_wb["Policy"], target_wb)
        _link_policy_values(target_wb["Policy"], target_wb)
        _copy_defined_names(source_wb, target_wb)
        target_wb.save(target)
    finally:
        source_wb.close()
        target_wb.close()

    return backup_path


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("target", type=Path, help="Workbook to patch in-place")
    parser.add_argument(
        "--module",
        type=Path,
        default=Path("CMO_Resin_Allocation_Module.xlsx"),
        help="Source workbook containing the Policy sheet",
    )
    parser.add_argument(
        "--no-backup",
        action="store_true",
        help="Do not create a sibling _BACKUP copy before modifying the target",
    )
    args = parser.parse_args()

    target_path = args.target
    module_path = args.module
    if not target_path.exists():
        raise FileNotFoundError(f"Target workbook not found: {target_path}")
    if not module_path.exists():
        raise FileNotFoundError(f"Module workbook not found: {module_path}")

    backup_path = inject_policy(target_path, module_path, backup=not args.no_backup)
    if backup_path != target_path:
        print(f"Backup written to {backup_path}")


if __name__ == "__main__":
    main()
