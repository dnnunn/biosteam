"""Helpers to snapshot key KPIs from the Excel baseline for regression tests."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, Mapping, Optional

import json
import math

from openpyxl import load_workbook
import yaml

__all__ = [
    "BaselineMetrics",
    "load_baseline_metrics",
    "export_baseline_metrics",
]


@dataclass(frozen=True)
class BaselineMetrics:
    """Container for the KPIs we care about for regression checks."""

    workbook_path: Path
    mass_trail: Mapping[str, float]
    final_product_kg: float
    cost_per_kg_usd: Optional[float] = None
    total_cost_per_batch_usd: Optional[float] = None
    cmo_fees_usd: Optional[float] = None
    materials_cost_per_batch_usd: Optional[float] = None
    materials_cost_breakdown: Mapping[str, float] = field(default_factory=dict)
    allocation_basis: Optional[str] = None
    allocation_denominator: Optional[float] = None
    cmo_per_unit_usd: Optional[float] = None
    resin_per_unit_usd: Optional[float] = None
    total_per_unit_usd: Optional[float] = None

    @property
    def product_per_batch_kg(self) -> float:
        return self.final_product_kg

    def as_dict(self) -> Dict[str, object]:
        return {
            "workbook": str(self.workbook_path),
            "mass_trail": dict(self.mass_trail),
            "final_product_kg": self.final_product_kg,
            "product_per_batch_kg": self.product_per_batch_kg,
            "cost_per_kg_usd": self.cost_per_kg_usd,
            "total_cost_per_batch_usd": self.total_cost_per_batch_usd,
            "cmo_fees_usd": self.cmo_fees_usd,
            "materials_cost_per_batch_usd": self.materials_cost_per_batch_usd,
            "materials_cost_breakdown": dict(self.materials_cost_breakdown),
            "allocation_basis": self.allocation_basis,
            "allocation_denominator": self.allocation_denominator,
            "cmo_per_unit_usd": self.cmo_per_unit_usd,
            "resin_per_unit_usd": self.resin_per_unit_usd,
            "total_per_unit_usd": self.total_per_unit_usd,
        }

    @classmethod
    def from_json(cls, path: Path | str) -> "BaselineMetrics":
        path = Path(path)
        with path.open("r", encoding="utf-8") as handle:
            data = json.load(handle)

        def _get_float(key: str) -> Optional[float]:
            value = data.get(key)
            if value is None:
                return None
            try:
                return float(value)
            except (TypeError, ValueError):
                return None

        return cls(
            workbook_path=Path(data.get("workbook", "")),
            mass_trail={k: float(v) for k, v in data.get("mass_trail", {}).items()},
            final_product_kg=float(data["final_product_kg"]),
            cost_per_kg_usd=_get_float("cost_per_kg_usd"),
            total_cost_per_batch_usd=_get_float("total_cost_per_batch_usd"),
            cmo_fees_usd=_get_float("cmo_fees_usd"),
            materials_cost_per_batch_usd=_get_float("materials_cost_per_batch_usd"),
            materials_cost_breakdown={
                k: float(v) for k, v in data.get("materials_cost_breakdown", {}).items()
            },
            allocation_basis=data.get("allocation_basis"),
            allocation_denominator=_get_float("allocation_denominator"),
            cmo_per_unit_usd=_get_float("cmo_per_unit_usd"),
            resin_per_unit_usd=_get_float("resin_per_unit_usd"),
            total_per_unit_usd=_get_float("total_per_unit_usd"),
        )


def export_baseline_metrics(
    *,
    workbook_path: Path | str = Path("BaselineModel.xlsx"),
    config_path: Path | str = Path("migration/baseline_metrics_map.yaml"),
    output_path: Path | str = Path("tests/opn/baseline_metrics.json"),
) -> BaselineMetrics:
    """Dump the baseline KPIs from ``workbook_path`` into ``output_path``."""

    metrics = load_baseline_metrics(workbook_path=workbook_path, config_path=config_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(metrics.as_dict(), f, indent=2, sort_keys=True)
    return metrics


def load_baseline_metrics(
    *,
    workbook_path: Path | str,
    config_path: Path | str = Path("migration/baseline_metrics_map.yaml"),
) -> BaselineMetrics:
    """Read KPI values from the Excel baseline according to ``config_path``."""

    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Baseline workbook not found: {workbook_path}")

    config = _load_config(config_path)

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        mass_trail = {}
        for name, descriptor in config["mass_trail"].items():
            value = _read_cell(descriptor, wb)
            if value is None:
                raise ValueError(
                    f"Failed to read mass-trail cell {descriptor!r} for {name}"
                )
            mass_trail[name] = value

        final_product_cell = config["final_product"]["cell"]
        final_product = _read_cell(final_product_cell, wb)
        if final_product is None:
            raise ValueError(
                f"Failed to read final product cell {final_product_cell!r}; update the map"
            )

        cost_per_kg_cell = config.get("cost_per_kg", {}).get("cell")
        cost_per_kg = _read_cell(cost_per_kg_cell, wb) if cost_per_kg_cell else None

        total_cost_cell = config.get("total_cost_per_batch", {}).get("cell")
        total_cost_per_batch = (
            _read_cell(total_cost_cell, wb) if total_cost_cell else None
        )

        cmo_fees_cell = config.get("cmo_fees", {}).get("cell")
        cmo_fees = _read_cell(cmo_fees_cell, wb) if cmo_fees_cell else None

        materials_cost_cell = config.get("materials_cost_per_batch", {}).get("cell")
        materials_cost_per_batch = (
            _read_cell(materials_cost_cell, wb) if materials_cost_cell else None
        )

        material_breakdown: Dict[str, float] = {}
        for key, descriptor in config.get("materials_costs", {}).items():
            value = _read_cell(descriptor, wb)
            if value is not None:
                material_breakdown[key] = value

        allocation_cfg = config.get("allocation", {}) if isinstance(config, Mapping) else {}
        allocation_basis = _read_text(allocation_cfg.get("basis"), wb)
        allocation_denominator = _read_cell(allocation_cfg.get("denominator"), wb)
        cmo_per_unit = _read_cell(allocation_cfg.get("cmo_per_unit"), wb)
        resin_per_unit = _read_cell(allocation_cfg.get("resin_per_unit"), wb)
        total_per_unit = _read_cell(allocation_cfg.get("total_per_unit"), wb)
        batches_executed = _read_cell(allocation_cfg.get("batches_executed"), wb)

        policy_metrics = _compute_policy_metrics(wb)

        def _fallback(key: str, existing: Optional[float]) -> Optional[float]:
            value = policy_metrics.get(key)
            return existing if existing is not None else value

        allocation_basis = allocation_basis or policy_metrics.get("basis")
        if isinstance(allocation_basis, str):
            allocation_basis = allocation_basis.strip() or None
            if allocation_basis:
                allocation_basis = allocation_basis.upper()
        allocation_denominator = allocation_denominator or policy_metrics.get("denominator")
        cmo_per_unit = _fallback("cmo_per_unit", cmo_per_unit)
        resin_per_unit = _fallback("resin_per_unit", resin_per_unit)
        total_per_unit = _fallback("total_per_unit", total_per_unit)
        batches_executed = batches_executed or policy_metrics.get("batches_executed")

        cmo_total = policy_metrics.get("cmo_total")
        resin_total = policy_metrics.get("resin_total")

        if cmo_total is None and cmo_per_unit is not None and allocation_denominator:
            cmo_total = cmo_per_unit * allocation_denominator
        if resin_total is None and resin_per_unit is not None and allocation_denominator:
            resin_total = resin_per_unit * allocation_denominator

        if cmo_fees is None and cmo_total is not None and batches_executed:
            if batches_executed:
                cmo_fees = cmo_total / batches_executed
    finally:
        wb.close()

    return BaselineMetrics(
        workbook_path=workbook_path,
        mass_trail=mass_trail,
        final_product_kg=final_product,
        cost_per_kg_usd=cost_per_kg,
        total_cost_per_batch_usd=total_cost_per_batch,
        cmo_fees_usd=cmo_fees,
        materials_cost_per_batch_usd=(
            materials_cost_per_batch
            if materials_cost_per_batch is not None
            else (
                total_cost_per_batch - cmo_fees
                if total_cost_per_batch is not None and cmo_fees is not None
                else None
            )
        ),
        materials_cost_breakdown=material_breakdown,
        allocation_basis=allocation_basis,
        allocation_denominator=allocation_denominator,
        cmo_per_unit_usd=cmo_per_unit,
        resin_per_unit_usd=resin_per_unit,
        total_per_unit_usd=total_per_unit,
    )


# ---------------------------------------------------------------------------
# Internal helpers


def _load_config(config_path: Path | str) -> Mapping[str, object]:
    config_path = Path(config_path)
    if not config_path.exists():
        raise FileNotFoundError(f"Baseline config not found: {config_path}")
    with config_path.open("r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
    _validate_config(data)
    return data


def _validate_config(data: Mapping[str, object]) -> None:
    required_sections = {"mass_trail", "final_product"}
    missing = required_sections - data.keys()
    if missing:
        raise ValueError(f"Baseline config missing keys: {sorted(missing)}")


def _read_cell(descriptor: Optional[str], workbook) -> Optional[float]:
    if not descriptor:
        return None

    descriptor = descriptor.strip()
    if not descriptor:
        return None

    if "!" not in descriptor:
        defined = workbook.defined_names.get(descriptor)
        if defined is None:
            return None
        destinations = list(defined.destinations)
        if not destinations:
            return None
        sheet_name, address = destinations[0]
        if not address:
            return None
        return _read_cell(f"{sheet_name}!{address}", workbook)

    sheet, cell = descriptor.split("!", 1)
    sheet = sheet.strip()
    cell = cell.strip()
    try:
        ws = workbook[sheet]
    except KeyError:
        return None
    try:
        raw = ws[cell].value
    except Exception:
        return None
    return _coerce_float(raw)


def _read_text(descriptor: Optional[str], workbook) -> Optional[str]:
    if not descriptor:
        return None

    descriptor = descriptor.strip()
    if not descriptor:
        return None

    if "!" not in descriptor:
        defined = workbook.defined_names.get(descriptor)
        if defined is None:
            return None
        destinations = list(defined.destinations)
        if not destinations:
            return None
        sheet_name, address = destinations[0]
        if not address:
            return None
        return _read_text(f"{sheet_name}!{address}", workbook)

    sheet, cell = descriptor.split("!", 1)
    sheet = sheet.strip()
    cell = cell.strip()
    try:
        ws = workbook[sheet]
    except KeyError:
        return None
    try:
        raw = ws[cell].value
    except Exception:
        return None
    if raw is None:
        return None
    return str(raw)


def _coerce_float(value: object | None) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    try:
        number = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    if math.isnan(number):
        return None
    return number


def _parse_cell_descriptor(descriptor: str) -> tuple[str, int, int]:
    """Parse strings like ``Calculations!B121`` into sheet / 0-based row/col."""

    if "!" not in descriptor:
        raise ValueError(f"Expected SHEET!CELL format, got {descriptor!r}")
    sheet, cell = descriptor.split("!", 1)
    cell = cell.strip().upper()
    col_letters = ""
    row_digits = ""
    for ch in cell:
        if ch.isalpha():
            col_letters += ch
        elif ch.isdigit():
            row_digits += ch
        else:
            raise ValueError(f"Unsupported character {ch!r} in cell {descriptor!r}")
    if not col_letters or not row_digits:
        raise ValueError(f"Invalid cell reference {descriptor!r}")
    col_index = _letters_to_index(col_letters)
    row_index = int(row_digits) - 1
    return sheet, row_index, col_index


def _letters_to_index(letters: str) -> int:
    total = 0
    for ch in letters:
        total = total * 26 + (ord(ch) - ord("A") + 1)
    return total - 1


def _compute_policy_metrics(workbook) -> Dict[str, float]:
    try:
        ws = workbook["Policy"]
    except KeyError:
        return {}

    def _num(cell: str) -> Optional[float]:
        try:
            return _coerce_float(ws[cell].value)
        except Exception:
            return None

    def _text(cell: str) -> Optional[str]:
        value = ws[cell].value
        if value is None:
            return None
        return str(value)

    basis = _text("D2")
    campaigns_planned = _num("D4") or 0.0
    batches_planned_per_campaign = _num("D5") or 0.0
    campaign_days = _num("D6") or 0.0
    process_hours_per_batch = _num("D7") or 0.0
    campaign_fee = _num("D8") or 0.0
    suite_fee = _num("D9") or 0.0
    suite_fee_per_day = _num("D10") or 0.0
    per_batch_fee = _num("D11") or 0.0
    retainer_fee = _num("D12") or 0.0
    batches_executed = _num("D13") or 0.0
    good_batches_released = _num("D14") or batches_executed
    avg_kg_per_batch = _num("D15") or 0.0
    total_kg_released = _num("D16")
    resin_cost_per_l = _num("D17") or 0.0
    resin_volume_l = _num("D18") or 0.0
    resin_salvage_fraction = _num("D19") or 0.0
    resin_lifetime_cycles = _num("D20") or 0.0
    cycles_per_batch = _num("D21") or 0.0
    cip_cost_per_cycle = _num("D22") or 0.0

    planned_batches = campaigns_planned * batches_planned_per_campaign

    cmo_total = (
        retainer_fee
        + campaigns_planned * (campaign_fee + suite_fee)
        + campaigns_planned * campaign_days * suite_fee_per_day
        + batches_executed * per_batch_fee
    )

    resin_gross_cost = resin_cost_per_l * resin_volume_l
    amortizable = resin_gross_cost * (1.0 - resin_salvage_fraction)
    amort_per_cycle = amortizable / resin_lifetime_cycles if resin_lifetime_cycles else 0.0
    total_cycles_used = batches_executed * cycles_per_batch
    resin_amort_used = amort_per_cycle * min(total_cycles_used, resin_lifetime_cycles or total_cycles_used)
    cip_total = total_cycles_used * cip_cost_per_cycle
    resin_total = resin_amort_used + cip_total

    if total_kg_released is None or total_kg_released <= 0.0:
        total_kg_released = good_batches_released * avg_kg_per_batch

    basis_key = (basis or "KG_RELEASED").strip().upper()
    if basis_key == "KG_RELEASED":
        denominator = total_kg_released
    elif basis_key == "GOOD_BATCHES":
        denominator = good_batches_released
    elif basis_key == "SCHEDULED_CAPACITY":
        denominator = planned_batches
    elif basis_key == "PROCESS_TIME_HOURS":
        denominator = batches_executed * process_hours_per_batch
    else:
        denominator = total_kg_released

    if denominator and denominator > 0.0:
        cmo_per_unit = cmo_total / denominator
        resin_per_unit = resin_total / denominator
        total_per_unit = (cmo_total + resin_total) / denominator
    else:
        cmo_per_unit = resin_per_unit = total_per_unit = None

    return {
        "basis": basis_key,
        "denominator": denominator,
        "cmo_total": cmo_total,
        "resin_total": resin_total,
        "cmo_per_unit": cmo_per_unit,
        "resin_per_unit": resin_per_unit,
        "total_per_unit": total_per_unit,
        "batches_executed": batches_executed,
        "total_kg_released": total_kg_released,
    }
