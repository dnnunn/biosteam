#!/usr/bin/env python3
"""
Enhanced Excel Analysis for Modular Framework
Analyzes the updated Excel model focusing on module framework, dropdown_lookup integration,
and alternative technologies like chitosan adsorption.
"""

import pandas as pd
import openpyxl
import json
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')

class ModularExcelAnalyzer:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.workbook = load_workbook(excel_path, data_only=True)
        self.worksheet_names = self.workbook.sheetnames
        self.module_data = {}
        self.dropdown_data = {}
        self.process_modules = {}

    def analyze_worksheet_structure(self):
        """Analyze the overall structure of worksheets"""
        print("=== WORKSHEET STRUCTURE ANALYSIS ===")
        print(f"Total worksheets: {len(self.worksheet_names)}")

        for i, name in enumerate(self.worksheet_names):
            ws = self.workbook[name]
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"{i+1:2d}. {name:<25} | Rows: {max_row:4d} | Cols: {max_col:3d}")

        return self.worksheet_names

    def analyze_dropdown_lookup(self):
        """Analyze the Dropdown_lookup worksheet for module definitions"""
        print("\n=== DROPDOWN_LOOKUP ANALYSIS ===")

        if 'Dropdown_lookup' not in self.worksheet_names:
            print("Dropdown_lookup worksheet not found")
            return None

        try:
            # Read the dropdown lookup sheet
            df = pd.read_excel(self.excel_path, sheet_name='Dropdown_lookup')
            print(f"Dropdown_lookup shape: {df.shape}")
            print("\nColumn headers:")
            for i, col in enumerate(df.columns):
                print(f"  {i+1}. {col}")

            # Look for module-related data
            print("\nFirst 10 rows of data:")
            print(df.head(10).to_string())

            # Save for further analysis
            self.dropdown_data = df

            # Look for patterns in the data
            if 'Module' in df.columns or any('module' in str(col).lower() for col in df.columns):
                print("\n--- MODULE-RELATED COLUMNS FOUND ---")
                module_cols = [col for col in df.columns if 'module' in str(col).lower()]
                for col in module_cols:
                    unique_vals = df[col].dropna().unique()
                    print(f"\n{col}: {len(unique_vals)} unique values")
                    print(f"Values: {unique_vals[:10]}")  # First 10 values

            return df

        except Exception as e:
            print(f"Error reading Dropdown_lookup: {e}")
            return None

    def analyze_inputs_assumptions(self):
        """Analyze the Inputs/Assumptions worksheet for module framework"""
        print("\n=== INPUTS/ASSUMPTIONS ANALYSIS ===")

        target_sheets = ['Inputs/Assumptions', 'Inputs', 'Assumptions']
        found_sheet = None

        for sheet in target_sheets:
            if sheet in self.worksheet_names:
                found_sheet = sheet
                break

        if not found_sheet:
            print("Inputs/Assumptions worksheet not found")
            print("Available sheets that might contain assumptions:")
            assumption_sheets = [s for s in self.worksheet_names if 'input' in s.lower() or 'assumption' in s.lower()]
            for sheet in assumption_sheets:
                print(f"  - {sheet}")
            return None

        try:
            df = pd.read_excel(self.excel_path, sheet_name=found_sheet)
            print(f"Found sheet: {found_sheet}")
            print(f"Shape: {df.shape}")

            # Look for module-related structure
            print("\nColumn headers:")
            for i, col in enumerate(df.columns):
                print(f"  {i+1}. {col}")

            # Search for module patterns in the data
            module_patterns = ['module', 'variant', 'alternative', 'option', 'technology']

            print("\n--- SEARCHING FOR MODULE PATTERNS ---")
            for pattern in module_patterns:
                matching_cols = [col for col in df.columns if pattern in str(col).lower()]
                if matching_cols:
                    print(f"\n{pattern.upper()} columns found:")
                    for col in matching_cols:
                        print(f"  - {col}")
                        unique_vals = df[col].dropna().unique()
                        print(f"    Unique values ({len(unique_vals)}): {unique_vals[:5]}")

            # Look for specific technology mentions
            chitosan_refs = []
            iex_refs = []

            for col in df.columns:
                col_data = df[col].dropna().astype(str)
                chitosan_matches = col_data[col_data.str.contains('chitosan', case=False, na=False)]
                iex_matches = col_data[col_data.str.contains('iex|ion exchange', case=False, na=False)]

                if len(chitosan_matches) > 0:
                    chitosan_refs.extend([(col, val) for val in chitosan_matches])
                if len(iex_matches) > 0:
                    iex_refs.extend([(col, val) for val in iex_matches])

            if chitosan_refs:
                print(f"\n--- CHITOSAN REFERENCES FOUND ({len(chitosan_refs)}) ---")
                for col, val in chitosan_refs[:5]:  # First 5
                    print(f"  {col}: {val}")

            if iex_refs:
                print(f"\n--- IEX REFERENCES FOUND ({len(iex_refs)}) ---")
                for col, val in iex_refs[:5]:  # First 5
                    print(f"  {col}: {val}")

            return df

        except Exception as e:
            print(f"Error reading {found_sheet}: {e}")
            return None

    def analyze_calculations_sheet(self):
        """Analyze the Calculations worksheet for module implementation"""
        print("\n=== CALCULATIONS SHEET ANALYSIS ===")

        calc_sheets = [s for s in self.worksheet_names if 'calc' in s.lower()]

        if not calc_sheets:
            print("No Calculations worksheet found")
            return None

        found_sheet = calc_sheets[0]
        print(f"Analyzing: {found_sheet}")

        try:
            df = pd.read_excel(self.excel_path, sheet_name=found_sheet)
            print(f"Shape: {df.shape}")

            # Look for modular calculation structure
            print("\nColumn headers:")
            for i, col in enumerate(df.columns[:20]):  # First 20 columns
                print(f"  {i+1:2d}. {col}")

            # Search for process modules and unit operations
            unit_op_keywords = [
                'fermentation', 'bioreactor', 'centrifuge', 'chromatography',
                'chitosan', 'adsorption', 'filtration', 'concentration',
                'precipitation', 'crystallization', 'drying', 'formulation'
            ]

            print("\n--- UNIT OPERATION PATTERNS ---")
            for keyword in unit_op_keywords:
                matching_cols = [col for col in df.columns if keyword in str(col).lower()]
                if matching_cols:
                    print(f"\n{keyword.upper()}:")
                    for col in matching_cols:
                        print(f"  - {col}")

            return df

        except Exception as e:
            print(f"Error reading {found_sheet}: {e}")
            return None

    def identify_process_modules(self):
        """Identify discrete process modules and their variants"""
        print("\n=== PROCESS MODULE IDENTIFICATION ===")

        # Read multiple sheets to build complete picture
        sheets_to_analyze = []

        # Priority sheets
        priority_sheets = ['Dropdown_lookup', 'Inputs/Assumptions', 'Calculations', 'Process', 'Modules']

        for sheet in priority_sheets:
            if sheet in self.worksheet_names:
                sheets_to_analyze.append(sheet)
            else:
                # Look for similar names
                similar = [s for s in self.worksheet_names if sheet.lower() in s.lower()]
                sheets_to_analyze.extend(similar)

        print(f"Analyzing sheets: {sheets_to_analyze}")

        # Extract module data from each sheet
        all_module_data = {}

        for sheet_name in sheets_to_analyze:
            try:
                df = pd.read_excel(self.excel_path, sheet_name=sheet_name)

                # Look for module structure
                module_info = {
                    'sheet': sheet_name,
                    'shape': df.shape,
                    'columns': list(df.columns),
                    'module_columns': [],
                    'technology_alternatives': [],
                    'cost_columns': []
                }

                # Identify module-related columns
                for col in df.columns:
                    col_str = str(col).lower()
                    if any(pattern in col_str for pattern in ['module', 'variant', 'alternative', 'option']):
                        module_info['module_columns'].append(col)

                    if any(pattern in col_str for pattern in ['cost', 'price', 'capex', 'opex']):
                        module_info['cost_columns'].append(col)

                # Look for technology alternatives in data
                for col in df.columns:
                    try:
                        col_data = df[col].dropna().astype(str)
                        tech_alternatives = []

                        for tech in ['chitosan', 'iex', 'ion exchange', 'chromatography', 'adsorption']:
                            matches = col_data[col_data.str.contains(tech, case=False, na=False)]
                            if len(matches) > 0:
                                tech_alternatives.extend(matches.tolist())

                        if tech_alternatives:
                            module_info['technology_alternatives'].extend(tech_alternatives)
                    except:
                        continue

                all_module_data[sheet_name] = module_info

            except Exception as e:
                print(f"Error analyzing {sheet_name}: {e}")
                continue

        self.process_modules = all_module_data
        return all_module_data

    def analyze_chitosan_vs_iex(self):
        """Detailed analysis of chitosan adsorption vs IEX chromatography"""
        print("\n=== CHITOSAN vs IEX ANALYSIS ===")

        chitosan_data = {}
        iex_data = {}

        # Search across all sheets for chitosan and IEX references
        for sheet_name in self.worksheet_names:
            try:
                df = pd.read_excel(self.excel_path, sheet_name=sheet_name)

                # Search for chitosan references
                for col in df.columns:
                    try:
                        col_data = df[col].dropna().astype(str)

                        # Chitosan references
                        chitosan_mask = col_data.str.contains('chitosan', case=False, na=False)
                        if chitosan_mask.any():
                            chitosan_matches = col_data[chitosan_mask]
                            if sheet_name not in chitosan_data:
                                chitosan_data[sheet_name] = {}
                            chitosan_data[sheet_name][col] = chitosan_matches.tolist()

                        # IEX references
                        iex_mask = col_data.str.contains('iex|ion exchange|chromatography', case=False, na=False)
                        if iex_mask.any():
                            iex_matches = col_data[iex_mask]
                            if sheet_name not in iex_data:
                                iex_data[sheet_name] = {}
                            iex_data[sheet_name][col] = iex_matches.tolist()

                    except:
                        continue

            except Exception as e:
                continue

        # Print findings
        print("\n--- CHITOSAN ADSORPTION REFERENCES ---")
        for sheet, cols in chitosan_data.items():
            print(f"\n{sheet}:")
            for col, matches in cols.items():
                print(f"  {col}: {len(matches)} references")
                for match in matches[:3]:  # First 3
                    print(f"    - {match}")

        print("\n--- IEX CHROMATOGRAPHY REFERENCES ---")
        for sheet, cols in iex_data.items():
            print(f"\n{sheet}:")
            for col, matches in cols.items():
                print(f"  {col}: {len(matches)} references")
                for match in matches[:3]:  # First 3
                    print(f"    - {match}")

        return {'chitosan': chitosan_data, 'iex': iex_data}

    def extract_biosteam_mapping(self):
        """Extract mapping framework for BioSTEAM unit operations"""
        print("\n=== BIOSTEAM MAPPING FRAMEWORK ===")

        # Define potential BioSTEAM unit operation mappings
        biosteam_mappings = {
            'upstream': {
                'seed_train': 'biosteam.units.BatchReactor',
                'fermentation': 'biosteam.units.BatchReactor',
                'media_prep': 'biosteam.units.Mixer'
            },
            'downstream': {
                'cell_separation': 'biosteam.units.SolidsCentrifuge',
                'chitosan_adsorption': 'biosteam.units.AdsorptionColumnTSA',
                'iex_chromatography': 'biosteam.units.MultiStageEquilibrium',
                'concentration': 'biosteam.units.MembraneBioreactor',
                'crystallization': 'biosteam.units.CrystallizerNaCl',
                'drying': 'biosteam.units.DrumDryer',
                'formulation': 'biosteam.units.Mixer'
            },
            'utilities': {
                'chilled_water': 'biosteam.facilities.ChilledWaterPackage',
                'steam': 'biosteam.facilities.BoilerTurbogenerator',
                'cooling_tower': 'biosteam.facilities.CoolingTower'
            }
        }

        print("Proposed BioSTEAM Unit Operation Mappings:")
        for category, mappings in biosteam_mappings.items():
            print(f"\n{category.upper()}:")
            for process, unit in mappings.items():
                print(f"  {process:<20} → {unit}")

        return biosteam_mappings

    def generate_comprehensive_report(self):
        """Generate comprehensive analysis report"""
        print("\n" + "="*80)
        print("COMPREHENSIVE MODULAR FRAMEWORK ANALYSIS")
        print("="*80)

        # Run all analyses
        self.analyze_worksheet_structure()
        dropdown_data = self.analyze_dropdown_lookup()
        inputs_data = self.analyze_inputs_assumptions()
        calc_data = self.analyze_calculations_sheet()
        module_data = self.identify_process_modules()
        tech_comparison = self.analyze_chitosan_vs_iex()
        biosteam_mapping = self.extract_biosteam_mapping()

        # Generate summary
        summary = {
            'worksheets': self.worksheet_names,
            'module_framework': module_data,
            'technology_alternatives': tech_comparison,
            'biosteam_mappings': biosteam_mapping,
            'analysis_timestamp': pd.Timestamp.now().isoformat()
        }

        return summary

def main():
    excel_path = "/Users/davidnunn/Desktop/Apps/BetterDairy/TEAM/Revised Model_15052025v44.xlsx"

    print("Initializing Modular Excel Analysis...")
    print(f"Target file: {excel_path}")

    try:
        analyzer = ModularExcelAnalyzer(excel_path)
        report = analyzer.generate_comprehensive_report()

        # Save detailed analysis
        output_path = "/Users/davidnunn/Desktop/Apps/BetterDairy/Biosteam/modular_framework_analysis.json"
        with open(output_path, 'w') as f:
            json.dump(report, f, indent=2, default=str)

        print(f"\n--- ANALYSIS COMPLETE ---")
        print(f"Detailed report saved to: {output_path}")

    except Exception as e:
        print(f"Analysis failed: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()