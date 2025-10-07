#!/usr/bin/env python3
"""
Excel Parameter Extractor for BioSTEAM Migration
Extracts all parameters from the osteopontin production Excel model
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
import json
import re
from pathlib import Path

class ExcelParameterExtractor:
    def __init__(self, excel_path):
        self.excel_path = Path(excel_path)
        self.workbook = load_workbook(self.excel_path, data_only=True)
        self.parameters = {}
        self.process_flows = {}
        self.economics = {}
        self.unit_operations = {}
        self.gaps_identified = []

    def extract_all_parameters(self):
        """Extract all parameters from all worksheets"""
        print(f"Analyzing Excel file: {self.excel_path}")
        print(f"Available worksheets: {self.workbook.sheetnames}")

        # Analyze each worksheet
        for sheet_name in self.workbook.sheetnames:
            print(f"\nAnalyzing sheet: {sheet_name}")
            self.analyze_worksheet(sheet_name)

        # Extract specific parameter categories
        self.extract_process_parameters()
        self.extract_economic_parameters()
        self.extract_unit_operation_parameters()
        self.identify_model_gaps()

        return {
            'parameters': self.parameters,
            'process_flows': self.process_flows,
            'economics': self.economics,
            'unit_operations': self.unit_operations,
            'gaps_identified': self.gaps_identified
        }

    def analyze_worksheet(self, sheet_name):
        """Analyze individual worksheet for parameters"""
        ws = self.workbook[sheet_name]
        sheet_params = {}

        # Convert worksheet to DataFrame for easier analysis
        try:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name, header=None)

            # Look for parameter patterns
            for idx, row in df.iterrows():
                for col_idx, cell in enumerate(row):
                    if pd.notna(cell) and isinstance(cell, str):
                        # Look for parameter-like patterns
                        if any(keyword in cell.lower() for keyword in [
                            'titer', 'yield', 'recovery', 'efficiency', 'cost', 'price',
                            'volume', 'time', 'temperature', 'ph', 'rate', 'capacity',
                            'factor', 'ratio', 'concentration', 'mass', 'flow'
                        ]):
                            # Try to find associated value
                            value = self.find_associated_value(df, idx, col_idx)
                            if value is not None:
                                sheet_params[cell] = value

        except Exception as e:
            print(f"Error analyzing sheet {sheet_name}: {e}")

        self.parameters[sheet_name] = sheet_params

    def find_associated_value(self, df, row_idx, col_idx):
        """Find numerical value associated with a parameter label"""
        # Check adjacent cells for numerical values
        search_positions = [
            (row_idx, col_idx + 1),  # Right
            (row_idx, col_idx + 2),  # Two right
            (row_idx + 1, col_idx),  # Below
            (row_idx - 1, col_idx),  # Above
        ]

        for r, c in search_positions:
            try:
                if r < len(df) and c < len(df.columns):
                    value = df.iloc[r, c]
                    if pd.notna(value) and isinstance(value, (int, float)):
                        return value
            except:
                continue

        return None

    def extract_process_parameters(self):
        """Extract key process parameters for fermentation and downstream"""
        process_keywords = {
            'fermentation': ['titer', 'yield', 'productivity', 'cycle time', 'volume'],
            'separation': ['recovery', 'efficiency', 'centrifuge', 'filtration'],
            'purification': ['chromatography', 'resin', 'column', 'buffer', 'elution'],
            'concentration': ['ultrafiltration', 'concentration', 'membrane'],
            'formulation': ['drying', 'packaging', 'storage']
        }

        for category, keywords in process_keywords.items():
            self.process_flows[category] = {}

            for sheet_name, params in self.parameters.items():
                for param_name, value in params.items():
                    if any(keyword in param_name.lower() for keyword in keywords):
                        self.process_flows[category][param_name] = {
                            'value': value,
                            'source_sheet': sheet_name
                        }

    def extract_economic_parameters(self):
        """Extract economic and costing parameters"""
        economic_keywords = [
            'cost', 'price', 'capex', 'opex', 'margin', 'revenue', 'profit',
            'depreciation', 'interest', 'tax', 'discount', 'npv', 'irr',
            'payback', 'working capital', 'contingency'
        ]

        for sheet_name, params in self.parameters.items():
            for param_name, value in params.items():
                if any(keyword in param_name.lower() for keyword in economic_keywords):
                    self.economics[param_name] = {
                        'value': value,
                        'source_sheet': sheet_name
                    }

    def extract_unit_operation_parameters(self):
        """Extract unit operation specific parameters"""
        unit_ops = {
            'fermentor': ['volume', 'titer', 'yield', 'cycle time', 'productivity'],
            'centrifuge': ['capacity', 'efficiency', 'recovery', 'throughput'],
            'chromatography': ['resin volume', 'bed height', 'flow rate', 'recovery'],
            'ultrafiltration': ['membrane area', 'flux', 'concentration factor'],
            'dryer': ['capacity', 'energy', 'efficiency', 'time']
        }

        for unit_op, keywords in unit_ops.items():
            self.unit_operations[unit_op] = {}

            for sheet_name, params in self.parameters.items():
                for param_name, value in params.items():
                    if any(keyword in param_name.lower() for keyword in keywords):
                        if unit_op not in self.unit_operations:
                            self.unit_operations[unit_op] = {}
                        self.unit_operations[unit_op][param_name] = {
                            'value': value,
                            'source_sheet': sheet_name
                        }

    def identify_model_gaps(self):
        """Identify known gaps in the Excel model"""
        gap_checks = [
            {
                'category': 'IEX Chromatography Buffers',
                'missing_parameters': ['loading buffer volume', 'wash buffer volume', 'elution buffer volume'],
                'impact': 'Underestimated buffer costs and preparation requirements'
            },
            {
                'category': 'Alternative Cell Separation',
                'missing_parameters': ['depth filtration', 'microfiltration options', 'membrane costs'],
                'impact': 'Limited separation technology options'
            },
            {
                'category': 'Adsorption Operations',
                'missing_parameters': ['adsorbent types', 'isotherm parameters', 'regeneration costs'],
                'impact': 'Missing alternative purification methods'
            },
            {
                'category': 'Scale-Dependent Parameters',
                'missing_parameters': ['scale factors for utilities', 'labor scaling', 'equipment scaling laws'],
                'impact': 'Inaccurate scaling predictions'
            },
            {
                'category': 'Process Intensification',
                'missing_parameters': ['continuous processing options', 'integrated operations', 'debottlenecking'],
                'impact': 'Missing advanced process options'
            },
            {
                'category': 'Quality and Regulatory',
                'missing_parameters': ['QC testing costs', 'regulatory filing costs', 'validation expenses'],
                'impact': 'Underestimated compliance costs'
            }
        ]

        for gap in gap_checks:
            # Check if parameters exist in extracted data
            found_params = []
            for missing_param in gap['missing_parameters']:
                found = self.search_parameter(missing_param)
                if found:
                    found_params.append(found)

            if len(found_params) < len(gap['missing_parameters']) * 0.5:  # Less than 50% found
                self.gaps_identified.append({
                    'category': gap['category'],
                    'missing_count': len(gap['missing_parameters']) - len(found_params),
                    'found_params': found_params,
                    'impact': gap['impact']
                })

    def search_parameter(self, search_term):
        """Search for parameter across all extracted data"""
        for sheet_name, params in self.parameters.items():
            for param_name, value in params.items():
                if search_term.lower() in param_name.lower():
                    return {'param': param_name, 'value': value, 'sheet': sheet_name}
        return None

    def generate_summary_report(self):
        """Generate comprehensive summary of extracted parameters"""
        report = {
            'extraction_summary': {
                'total_sheets': len(self.workbook.sheetnames),
                'total_parameters': sum(len(params) for params in self.parameters.values()),
                'process_categories': len(self.process_flows),
                'economic_parameters': len(self.economics),
                'unit_operations': len(self.unit_operations),
                'identified_gaps': len(self.gaps_identified)
            },
            'key_process_parameters': self.get_key_parameters(),
            'economic_structure': self.get_economic_structure(),
            'scale_parameters': self.get_scale_parameters(),
            'gaps_analysis': self.gaps_identified
        }
        return report

    def get_key_parameters(self):
        """Extract most critical process parameters"""
        key_params = {}

        # Critical fermentation parameters
        fermentation_keys = ['titer', 'yield', 'productivity', 'cycle time']
        for key in fermentation_keys:
            found = self.search_parameter(key)
            if found:
                key_params[f'fermentation_{key}'] = found

        # Critical separation parameters
        separation_keys = ['recovery', 'efficiency']
        for key in separation_keys:
            found = self.search_parameter(key)
            if found:
                key_params[f'separation_{key}'] = found

        return key_params

    def get_economic_structure(self):
        """Extract economic model structure"""
        return {
            'cost_categories': list(self.economics.keys()),
            'total_parameters': len(self.economics)
        }

    def get_scale_parameters(self):
        """Extract scaling-related parameters"""
        scale_params = {}
        scale_keywords = ['scale', 'capacity', 'volume', 'throughput', 'factor']

        for keyword in scale_keywords:
            found = self.search_parameter(keyword)
            if found:
                scale_params[keyword] = found

        return scale_params

def main():
    """Main execution function"""
    excel_path = "/Users/davidnunn/Desktop/Apps/BetterDairy/TEAM/Revised Model_15052025v29.xlsx"

    extractor = ExcelParameterExtractor(excel_path)
    results = extractor.extract_all_parameters()
    report = extractor.generate_summary_report()

    # Save results
    output_path = "/Users/davidnunn/Desktop/Apps/BetterDairy/Biosteam/excel_extraction_results.json"
    with open(output_path, 'w') as f:
        json.dump({
            'extraction_results': results,
            'summary_report': report
        }, f, indent=2, default=str)

    print("\n" + "="*80)
    print("EXCEL PARAMETER EXTRACTION COMPLETE")
    print("="*80)

    print(f"\nExtraction Summary:")
    print(f"- Total sheets analyzed: {report['extraction_summary']['total_sheets']}")
    print(f"- Total parameters found: {report['extraction_summary']['total_parameters']}")
    print(f"- Process categories: {report['extraction_summary']['process_categories']}")
    print(f"- Economic parameters: {report['extraction_summary']['economic_parameters']}")
    print(f"- Unit operations: {report['extraction_summary']['unit_operations']}")
    print(f"- Identified gaps: {report['extraction_summary']['identified_gaps']}")

    print(f"\nKey Process Parameters Found:")
    for param, details in report['key_process_parameters'].items():
        print(f"- {param}: {details['value']} (from {details['sheet']})")

    print(f"\nCritical Gaps Identified:")
    for gap in report['gaps_analysis']:
        print(f"- {gap['category']}: {gap['missing_count']} missing parameters")
        print(f"  Impact: {gap['impact']}")

    print(f"\nResults saved to: {output_path}")

    return results, report

if __name__ == "__main__":
    results, report = main()